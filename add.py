from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# add a simple formula
ws["A1"] = "=SUM(1, 1)"
ws["A2"] = "=MUL(A1, 5)"
wb.save("formula.xlsx")